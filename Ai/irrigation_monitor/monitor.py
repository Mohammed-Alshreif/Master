import os
import time
import json
import requests
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import google.generativeai as genai

# ─────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────
FIREBASE_URL    = "https://esptest1-edcd8-default-rtdb.firebaseio.com"
FIREBASE_PATH   = "/apartments/flat1/rooms/room1"
GEMINI_API_KEY  = "AQ.Ab8RN6JR04QOUYpZH2byFN_u5_C0O8gOojdZFwnTPAoUEfDu3A"

# ─────────────────────────────────────────
# USER MACROS
# ─────────────────────────────────────────


BASE_DIR        = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE      = os.path.join(BASE_DIR, "irrigation_log.xlsx")
POLL_INTERVAL   = 50        # seconds between each read
ANALYSIS_HOURS  = 4      # hours between each Gemini analysis

# ─────────────────────────────────────────
# FIREBASE
# ─────────────────────────────────────────
def firebase_get(path):
    try:
        resp = requests.get(f"{FIREBASE_URL}{path}.json", timeout=10)
        resp.raise_for_status()
        return resp.json()
    except Exception as e:
        print(f"[Firebase] Error: {e}")
        return None

def firebase_put(path, data):
    try:
        resp = requests.put(f"{FIREBASE_URL}{path}.json", json=data, timeout=10)
        resp.raise_for_status()
        return True
    except Exception as e:
        print(f"[Firebase PUT] Error: {e}")
        return False

# ─────────────────────────────────────────
# Extract snapshot data flexibly for any number of sensors/devices
# ─────────────────────────────────────────
def extract_snapshot(data):
    """
    Automatically reads all sensors and devices without predefining them.
    Returns a dict containing each sensor/device name and its value.
    """
    snapshot = {}

    # ── Sensors (flexible - any number) ──
    sensors = data.get("sensors", {})
    for key, val in sensors.items():
        name = val.get("name", key)
        unit = val.get("unit", "")
        value = val.get("value", None)
        col_name = f"{name} ({unit})" if unit else name
        snapshot[col_name] = value

    # ── Devices (flexible - any number) ──
    devices = data.get("devices", {})
    for key, val in devices.items():
        name = val.get("name", key)
        if "value" in val:
            snapshot[f"{name}"] = val["value"]
        elif "status" in val:
            snapshot[f"{name}"] = val["status"]

    return snapshot


def parse_date(date_str):
    if not date_str:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(date_str, fmt).date()
        except Exception:
            continue
    return None


def get_season(now):
    month = now.month
    if month in (12, 1, 2):
        return "winter"
    if month in (3, 4, 5):
        return "spring"
    if month in (6, 7, 8):
        return "summer"
    return "autumn"


def estimate_growth_stage(age_days):
    if age_days is None:
        return "unknown"
    if age_days < 14:
        return "seedling"
    if age_days < 45:
        return "vegetative"
    if age_days < 90:
        return "developing"
    return "mature"


def extract_plant_info(data):
    plant = data.get("PLANT", {})
    return {
        "name": plant.get("NAME", ""),
        "land_area": plant.get("Land area", ""),
        "planting_date": plant.get("Planting date", ""),
        "soil_type": plant.get("Soil type", ""),
        "soil_fertilizer": plant.get("Soil Fertilizer", plant.get("Soil fertilizer", "")),
        "last_fertilization": plant.get("Last fertilization", ""),
    }


def extract_device_info(data):
    devices = data.get("devices", {})
    fertilizer = {"key": None, "name": None, "status": None, "type": None}
    soil_moisture_thresholds = {"key": None, "min": None, "max": None}
    irrigation = {"key": None, "name": None, "status": None, "type": None, "last_irrigation": None}

    for key, val in devices.items():
        name = val.get("name", "")
        if key == "device1" or name == "Fertilizer Pump🌱":
            fertilizer = {
                "key": key,
                "name": name,
                "status": val.get("status"),
                "type": val.get("type"),
            }
        if key == "device3" or "Water Pump" in name:
            soil_moisture_thresholds = {
                "key": key,
                "min": val.get("MIN_HUM"),
                "max": val.get("Max_HUM"),
            }
            irrigation = {
                "key": key,
                "name": name,
                "status": val.get("status"),
                "type": val.get("type"),
                "last_irrigation": val.get("Last irrigation"),
            }

    return fertilizer, soil_moisture_thresholds, irrigation

# ─────────────────────────────────────────
# EXCEL
# ─────────────────────────────────────────
HEADER_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
DATA_FONT   = Font(name="Arial", size=10)
ALT_FILL    = PatternFill("solid", start_color="D9E1F2")
THIN        = Side(style="thin", color="B0B0B0")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header(cell):
    cell.font      = HEADER_FONT
    cell.fill      = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = BORDER

def style_data(cell, row_idx):
    cell.font      = DATA_FONT
    cell.border    = BORDER
    cell.alignment = Alignment(horizontal="center")
    if row_idx % 2 == 0:
        cell.fill = ALT_FILL


def safe_save(wb):
    try:
        wb.save(EXCEL_FILE)
        return True
    except PermissionError:
        print(f"[Excel] Error: Cannot save '{EXCEL_FILE}'. Make sure the file is not open in Excel.")
        return False
    except Exception as e:
        print(f"[Excel] Unexpected save error: {e}")
        return False


def init_excel(columns):
    """
    Creates or loads the Excel workbook.
    columns = list of column names extracted from Firebase.
    """
    headers_log = ["Timestamp"] + columns + ["Changed Fields"]

    if os.path.exists(EXCEL_FILE):
        wb         = load_workbook(EXCEL_FILE)
        ws_log     = wb["Sensor Log"]
        if "Period Averages" in wb.sheetnames:
            ws_avg = wb["Period Averages"]
        else:
            ws_avg = wb["4h Averages"]
        ws_gemini  = wb["Gemini Feedback"]
        if "Fertilization Schedule" in wb.sheetnames:
            ws_schedule = wb["Fertilization Schedule"]
        else:
            ws_schedule = wb.create_sheet("Fertilization Schedule")
            schedule_headers = [
                "Timestamp",
                "Period",
                "Event #",
                "Date",
                "Type",
                "Amount",
                "Note",
            ]
            for col_idx, h in enumerate(schedule_headers, 1):
                cell = ws_schedule.cell(row=1, column=col_idx, value=h)
                style_header(cell)
                ws_schedule.column_dimensions[get_column_letter(col_idx)].width = 25
            ws_schedule.freeze_panes = "A2"
    else:
        wb = Workbook()

        # Sheet 1: Sensor Log
        ws_log = wb.active
        ws_log.title = "Sensor Log"
        for col_idx, h in enumerate(headers_log, 1):
            cell = ws_log.cell(row=1, column=col_idx, value=h)
            style_header(cell)
            ws_log.column_dimensions[get_column_letter(col_idx)].width = 22
        ws_log.row_dimensions[1].height = 30
        ws_log.freeze_panes = "A2"

        # Sheet 2: Period Averages
        ws_avg = wb.create_sheet("Period Averages")
        avg_headers = ["Period Start", "Period End", "Sample Count"] + columns
        for col_idx, h in enumerate(avg_headers, 1):
            cell = ws_avg.cell(row=1, column=col_idx, value=h)
            style_header(cell)
            ws_avg.column_dimensions[get_column_letter(col_idx)].width = 22
        ws_avg.freeze_panes = "A2"

        # Sheet 3: Gemini Feedback
        ws_gemini = wb.create_sheet("Gemini Feedback")
        for col_idx, h in enumerate(
            ["Timestamp", "Period", "Gemini Recommendation", "Action Sent", "Status"], 1
        ):
            cell = ws_gemini.cell(row=1, column=col_idx, value=h)
            style_header(cell)
        ws_gemini.column_dimensions["A"].width = 22
        ws_gemini.column_dimensions["B"].width = 35
        ws_gemini.column_dimensions["C"].width = 70
        ws_gemini.column_dimensions["D"].width = 40
        ws_gemini.column_dimensions["E"].width = 15
        ws_gemini.freeze_panes = "A2"

        # Sheet 4: Fertilization Schedule
        ws_schedule = wb.create_sheet("Fertilization Schedule")
        schedule_headers = [
            "Timestamp",
            "Period",
            "Event #",
            "Date",
            "Type",
            "Amount",
            "Note",
        ]
        for col_idx, h in enumerate(schedule_headers, 1):
            cell = ws_schedule.cell(row=1, column=col_idx, value=h)
            style_header(cell)
            ws_schedule.column_dimensions[get_column_letter(col_idx)].width = 25
        ws_schedule.freeze_panes = "A2"

        safe_save(wb)
        print(f"[Excel] Created '{EXCEL_FILE}'")

    return wb, ws_log, ws_avg, ws_gemini, ws_schedule, headers_log

def log_row(wb, ws, headers_log, snapshot, changed, timestamp):
    row_idx = ws.max_row + 1
    row_data = [timestamp.strftime("%Y-%m-%d %H:%M:%S")]
    for col in headers_log[1:-1]:   # between Timestamp and Changed Fields
        row_data.append(snapshot.get(col, ""))
    row_data.append(changed)

    ws.append(row_data)
    for col_idx in range(1, len(row_data) + 1):
        style_data(ws.cell(row=row_idx, column=col_idx), row_idx)
    safe_save(wb)

def log_average(wb, ws_avg, period_start, period_end, buffer, columns):
    if not buffer:
        return {}

    avgs = {}
    for col in columns:
        vals = [r[col] for r in buffer if isinstance(r.get(col), (int, float))]
        avgs[col] = round(sum(vals) / len(vals), 2) if vals else None

    row_idx = ws_avg.max_row + 1
    row_data = [
        period_start.strftime("%Y-%m-%d %H:%M"),
        period_end.strftime("%Y-%m-%d %H:%M"),
        len(buffer)
    ] + [avgs.get(c) for c in columns]

    ws_avg.append(row_data)
    for col_idx in range(1, len(row_data) + 1):
        style_data(ws_avg.cell(row=row_idx, column=col_idx), row_idx)
    safe_save(wb)
    return avgs

def log_gemini(wb, ws_gemini, timestamp, period_str, recommendation, action_str, status):
    row_idx = ws_gemini.max_row + 1
    ws_gemini.append([
        timestamp.strftime("%Y-%m-%d %H:%M:%S"),
        period_str, recommendation, action_str, status
    ])
    for col_idx in range(1, 6):
        cell = ws_gemini.cell(row=row_idx, column=col_idx)
        cell.font      = DATA_FONT
        cell.border    = BORDER
        cell.alignment = Alignment(
            horizontal="left" if col_idx == 3 else "center", wrap_text=True
        )
        if row_idx % 2 == 0:
            cell.fill = ALT_FILL
    ws_gemini.row_dimensions[row_idx].height = 80
    safe_save(wb)


def log_schedule(wb, ws_schedule, timestamp, period_str, schedule):
    if not schedule:
        return

    for idx, item in enumerate(schedule, start=1):
        date = item.get("date") if isinstance(item, dict) else ""
        type_ = item.get("type") if isinstance(item, dict) else ""
        amount = item.get("amount") if isinstance(item, dict) else ""
        note = item.get("note") if isinstance(item, dict) else str(item)
        row_idx = ws_schedule.max_row + 1
        ws_schedule.append([
            timestamp.strftime("%Y-%m-%d %H:%M:%S"),
            period_str,
            idx,
            date,
            type_,
            amount,
            note,
        ])
        for col_idx in range(1, 8):
            cell = ws_schedule.cell(row=row_idx, column=col_idx)
            cell.font      = DATA_FONT
            cell.border    = BORDER
            cell.alignment = Alignment(horizontal="left", wrap_text=True)
            if row_idx % 2 == 0:
                cell.fill = ALT_FILL
    safe_save(wb)


def normalize_json_text(text):
    text = text.strip()
    if not text:
        return ""
    if text.startswith("```"):
        text = text.strip("`\n")
    if text.lower().startswith("json"):
        text = text[text.find("{") :].strip()
    start = text.find("{")
    end = text.rfind("}")
    if start != -1 and end != -1:
        text = text[start:end+1]
    return text

# ─────────────────────────────────────────
# GEMINI
# ─────────────────────────────────────────
def ask_gemini(avgs, period_str, period_end, plant_info, humidity_range, fertilizer_info, irrigation_info, season, growth_stage):
    genai.configure(api_key=GEMINI_API_KEY)
    model = genai.GenerativeModel("gemini-2.5-flash")

    # Separate soil moisture from ambient humidity in sensor data
    sensor_lines = ""
    soil_moisture_value = None
    ambient_humidity_value = None
    other_sensors = []
    
    for k, v in avgs.items():
        if v is not None:
            k_lower = k.lower()
            if "soil moisture" in k_lower or "soil water" in k_lower:
                soil_moisture_value = f"- {k}: {v}%"
            elif "ambient humidity" in k_lower or "air humidity" in k_lower or "rh" in k_lower:
                ambient_humidity_value = f"- {k}: {v}%"
            else:
                other_sensors.append(f"- {k}: {v}")
    
    # Build sensor section with explicit order and clarity
    if soil_moisture_value:
        sensor_lines += soil_moisture_value + "\n"
    if ambient_humidity_value:
        sensor_lines += ambient_humidity_value + "\n"
    sensor_lines += "\n".join(other_sensors)

    plant_lines = "\n".join(
        f"- {label}: {value}"
        for label, value in [
            ("Plant name", plant_info.get("name")),
            ("Land area", plant_info.get("land_area")),
            ("Planting date", plant_info.get("planting_date")),
            ("Soil type", plant_info.get("soil_type")),
            ("Soil fertilizer", plant_info.get("soil_fertilizer")),
            ("Last fertilization", plant_info.get("last_fertilization")),
            ("Estimated growth stage", growth_stage),
            ("Season", season),
        ]
        if value
    )

    soil_moisture_text = (
        f"Current allowed SOIL MOISTURE range: {humidity_range.get('min')} to {humidity_range.get('max')}% for irrigation control."
        if humidity_range.get("min") is not None and humidity_range.get("max") is not None
        else "Soil moisture thresholds are not available from Firebase."
    )

    irrigation_text = (
        f"{irrigation_info.get('name')} status={irrigation_info.get('status')}, last irrigation={irrigation_info.get('last_irrigation')}"
        if irrigation_info.get("name")
        else "Irrigation device information is unavailable."
    )

    prompt = f"""
You are an expert horticulture and irrigation advisor for indoor plant care.
Use the plant metadata, device settings, sensor averages, and current date to decide the best next actions.
Crops should be cared for according to FAO, IFA, and USDA fertilization recommendations.

*** CRITICAL SOIL MOISTURE THRESHOLDS REQUIREMENT ***
You MUST return humidity_min and humidity_max in the actions object. These are NOT optional.
humidity_min = minimum soil moisture (%) before irrigation activates — REQUIRED
humidity_max = maximum soil moisture (%) when irrigation should stop — REQUIRED
These are the MOST IMPORTANT values for controlling irrigation. Do NOT leave them empty or null.

Plant metadata:
{plant_lines}

Control devices:
- Fertilizer pump: {fertilizer_info.get('name')} (status={fertilizer_info.get('status')}, type={fertilizer_info.get('type')})
- Irrigation pump: {irrigation_text}
- Soil moisture thresholds (MIN_HUM, Max_HUM): {soil_moisture_text}

Current period sensor averages (CLEARLY LABELED):
{sensor_lines}

Data collection completed at: {period_end.strftime('%Y-%m-%d %H:%M')} (local time).

Important (must follow exactly):
- Calculate plant age using the planting date and the current date, and determine the growth stage.
- Consider the time of day when evaluating light (lux): if data collection time is during night hours (e.g. between 19:00 and 06:00 local time), do NOT treat low lux as a problem unless grow-lights are expected; if daytime, flag low lux and recommend supplemental lighting.
- *** SOIL MOISTURE CONTROL ***: Gemini MUST set humidity_min and humidity_max as SOIL MOISTURE thresholds (percentage), NOT ambient air humidity. These directly control irrigation: if soil moisture drops BELOW humidity_min, irrigation should activate; if it reaches humidity_max, irrigation should stop.
- If current soil moisture is below the computed humidity_min, IMMEDIATELY recommend irrigation (set irrigation_pump_status=true and update last_irrigation timestamp).
- Compute humidity_min and humidity_max based on plant stage, soil type, and FAO/IFA/USDA irrigation guidelines. For example: seedling on peat may need 30-70% soil moisture range, while mature plants might be 20-60%.
- MUST return the exact in this time and the next time `fertilizer_type` and `fertilizer_timing` fields in the `actions` object according to FAO, IFA, and USDA recommendations table. These are REQUIRED: if you recommend 'no fertilization', return `fertilizer_type": "none"` and an appropriate `fertilizer_timing` like "none" or "N/A" in this time and the next time write the fertilization table in the next time.

Instructions:
1. Give a 1-2 sentence assessment of plant history and how to care it. 
2. Give a 1-2 sentence assessment of plant health and soil moisture state. 
3. Return a short recommendation (1-2 sentences) for soil moisture management and irrigation strategy.
4. **CALCULATE and RETURN soil moisture thresholds with FULL JUSTIFICATION**:
   - Cite the specific FAO, IFA, or USDA guidelines you are using (e.g., "FAO irrigation manual for herbs recommends...")
   - Explain the REASON for each value (e.g., "humidity_min=25% because seedlings on peat need frequent watering to prevent wilting; humidity_max=65% to avoid root rot...")
   - Return exact numeric `humidity_min` and `humidity_max` values — THESE ARE REQUIRED. Do NOT return null or empty values.
   - Example response: "Based on FAO guidelines for leafy herbs, humidity_min=25% (seedling stage requires frequent moisture) and humidity_max=65% (prevents waterlogging on peat soil)"
5. MUST build a `fertilization_schedule`: a list of 2-3 upcoming fertilization events for this crop, each with an exact or relative date, the fertilizer type/dose, and a short reason, based on FAO/IFA/USDA guidance for this growth stage.


Respond ONLY with valid JSON matching this schema:
{{
    "assessment": "...",
    "recommendation": "...",
    "soil_moisture_rationale": "Brief explanation citing FAO/IFA/USDA guidelines and the reason for chosen humidity_min and humidity_max values",
    "actions": {{
        "fertilizer_pump_status": true,
        "plant_last_fertilization": "YYYY-MM-DD",
        "fertilizer_type": "...",
        "fertilizer_timing": "YYYY-MM-DD or 'in N days' or 'none'",
        "irrigation_pump_status": true,
        "last_irrigation": "YYYY-MM-DD HH:MM",
        "humidity_min": 0,
        "humidity_max": 100
    }},
    "fertilization_schedule": [
        {{"date": "YYYY-MM-DD or 'today'", "type": "e.g. compost / NPK 5-5-5", "amount": "e.g. 1 tsp diluted in 1L water", "note": "why this application"}}
    ]
}}
"""
    try:
        response = model.generate_content(prompt)
        text = normalize_json_text(response.text)
        if not text:
            raise ValueError("Empty Gemini response")
        parsed = json.loads(text)
        recommendation = f"{parsed.get('assessment','')} | {parsed.get('recommendation','')}"
        soil_moisture_rationale = parsed.get('soil_moisture_rationale', 'No rationale provided')
        actions = parsed.get("actions", {})
        schedule = parsed.get("fertilization_schedule", [])
        if not isinstance(schedule, list):
            schedule = [schedule]
        # Ensure mandatory action fields exist; fill sensible defaults when missing
        soil = plant_info.get("soil_type", "").lower() if plant_info else ""
        if not actions.get("fertilizer_type"):
            if "peat" in soil or "moss" in soil or "sand" in soil:
                actions["fertilizer_type"] = "compost (light)"
            else:
                actions["fertilizer_type"] = "balanced NPK"

        if not actions.get("fertilizer_timing"):
            if growth_stage == "seedling":
                actions["fertilizer_timing"] = "in 7 days"
            else:
                actions["fertilizer_timing"] = "in 14 days"

        # Ensure humidity numeric bounds
        try:
            if actions.get("humidity_min") is None:
                actions["humidity_min"] = max(30, int(humidity_range.get("min") or 50))
            if actions.get("humidity_max") is None:
                actions["humidity_max"] = min(90, int(humidity_range.get("max") or 70))
        except Exception:
            actions.setdefault("humidity_min", 50)
            actions.setdefault("humidity_max", 70)

        return recommendation, soil_moisture_rationale, actions, schedule
    except Exception as e:
        raw_text = getattr(response, 'text', '<no response>') if 'response' in locals() else '<no response>'
        print(f"[Gemini] Error: {e}")
        print(f"[Gemini] Raw response: {repr(raw_text)}")
        soil_moisture_rationale = f"Error occurred, using defaults. Seedling needs frequent watering. Based on FAO guidelines: humidity_min=40%, humidity_max=70%"
        fallback = {
            "fertilizer_pump_status": False,
            "plant_last_fertilization": plant_info.get("last_fertilization", ""),
            "fertilizer_type": "NPK blend or compost based on soil",
            "fertilizer_timing": "in 7 days",
            "irrigation_pump_status": False,
            "last_irrigation": irrigation_info.get("last_irrigation", ""),
            "humidity_min": 40,
            "humidity_max": 70,
        }
        fallback_schedule = [
            {
                "date": "in 7 days",
                "type": "Compost diluted",
                "amount": "1 tsp compost in 1L water",
                "note": "Light feeding for seedling stage",
            }
        ]
        return (
            f"Error: {e}",
            soil_moisture_rationale,
            fallback,
            fallback_schedule,
        )

def apply_actions(actions, fertilizer_info, soil_moisture_thresholds, irrigation_info):
    """Send Gemini decisions to Firebase automatically for the defined devices."""
    success = True

    if fertilizer_info.get("key") and actions.get("fertilizer_pump_status") is not None:
        ok = firebase_put(
            f"{FIREBASE_PATH}/devices/{fertilizer_info['key']}/status",
            actions["fertilizer_pump_status"],
        )
        success = success and ok

    if actions.get("plant_last_fertilization"):
        ok = firebase_put(
            f"{FIREBASE_PATH}/PLANT/Last fertilization",
            actions["plant_last_fertilization"],
        )
        success = success and ok

    if irrigation_info.get("key") and actions.get("irrigation_pump_status") is not None:
        ok = firebase_put(
            f"{FIREBASE_PATH}/devices/{irrigation_info['key']}/status",
            actions["irrigation_pump_status"],
        )
        success = success and ok
        if actions.get("last_irrigation"):
            ok2 = firebase_put(
                f"{FIREBASE_PATH}/devices/{irrigation_info['key']}/Last irrigation",
                actions["last_irrigation"],
            )
            success = success and ok2

    if soil_moisture_thresholds.get("key"):
        if actions.get("humidity_min") is not None:
            ok = firebase_put(
                f"{FIREBASE_PATH}/devices/{soil_moisture_thresholds['key']}/MIN_HUM",
                actions["humidity_min"],
            )
            success = success and ok
        if actions.get("humidity_max") is not None:
            ok = firebase_put(
                f"{FIREBASE_PATH}/devices/{soil_moisture_thresholds['key']}/Max_HUM",
                actions["humidity_max"],
            )
            success = success and ok

    return success
def print_fertilization_table(schedule):
    print("[Fertilization Schedule]")
    print(f"{'Date':<15}{'Type':<20}{'Amount':<25}{'Note'}")
    print("-" * 80)
    for entry in schedule:
        print(f"{entry.get('date',''):<15}{entry.get('type',''):<20}{entry.get('amount',''):<25}{entry.get('note','')}")
# ─────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────
def main():
    print("=" * 50)
    print("  Irrigation Monitor - Starting")
    print("=" * 50)

    # First read to determine data structure
    print("[Firebase] Reading initial data...")
    data = firebase_get(FIREBASE_PATH)
    if not data:
        print("[Error] Cannot connect to Firebase. Check your connection.")
        return

    first_snapshot = extract_snapshot(data)
    columns = list(first_snapshot.keys())
    print(f"[Firebase] Detected {len(columns)} columns automatically:")
    for c in columns:
        print(f"    • {c}")

    wb, ws_log, ws_avg, ws_gemini, ws_schedule, headers_log = init_excel(columns)

    prev_snapshot  = {}
    buffer         = []
    period_start   = datetime.now()
    next_analysis  = datetime.now() + timedelta(hours=ANALYSIS_HOURS)

    print(f"\n[Monitor] Monitoring every {POLL_INTERVAL} seconds...")
    print(f"[Monitor] Next analysis: {next_analysis.strftime('%H:%M:%S')}\n")

    plant_info = extract_plant_info(data)
    fertilizer_info, soil_moisture_thresholds, irrigation_info = extract_device_info(data)
    planting_date = parse_date(plant_info.get("planting_date"))
    plant_age_days = (datetime.now().date() - planting_date).days if planting_date else None
    growth_stage = estimate_growth_stage(plant_age_days)
    season = get_season(datetime.now())

    while True:
        now  = datetime.now()
        data = firebase_get(FIREBASE_PATH)

        if data:
            snapshot = extract_snapshot(data)
            plant_info = extract_plant_info(data)
            fertilizer_info, soil_moisture_thresholds, irrigation_info = extract_device_info(data)

            # Detect changes
            changed_fields = [
                f"{k}: {prev_snapshot.get(k)} → {snapshot.get(k)}"
                for k in snapshot
                if prev_snapshot.get(k) != snapshot.get(k)
            ]
            changed_str = " | ".join(changed_fields) if changed_fields else ""

            if changed_str or not prev_snapshot:
                log_row(wb, ws_log, headers_log, snapshot, changed_str or "initial", now)
                print(f"[{now.strftime('%H:%M:%S')}] {changed_str or 'initial read'}")

            prev_snapshot = snapshot.copy()
            buffer.append(snapshot.copy())

        # ── Analysis interval check ──────────────────
        if now >= next_analysis:
            print(f"\n[Analysis] Analyzing {len(buffer)} samples...")
            avgs       = log_average(wb, ws_avg, period_start, now, buffer, columns)
            period_str = (f"{period_start.strftime('%Y-%m-%d %H:%M')} → "
                          f"{now.strftime('%Y-%m-%d %H:%M')}")

            recommendation, soil_moisture_rationale, actions, schedule = ask_gemini(
                avgs,
                period_str,
                now,
                plant_info,
                soil_moisture_thresholds,
                fertilizer_info,
                irrigation_info,
                season,
                growth_stage,
            )
            print("=" * 50)
            print("  AI Recommendation  ")
            print("=" * 50)
            print(f"[Gemini] {recommendation}")
            print(f"\n[Soil Moisture Thresholds Justification]")
            print(f"  {soil_moisture_rationale}")
            print(f"\n[Soil Moisture Thresholds]")
            print(f"  MIN (irrigation starts): {actions.get('humidity_min', 'N/A')}%")
            print(f"  MAX (irrigation stops):  {actions.get('humidity_max', 'N/A')}%")
            print_fertilization_table(schedule)
            print("[Gemini] Raw fertilization schedule:")
            print(json.dumps(schedule, indent=2, ensure_ascii=False))
            fb_ok      = apply_actions(actions, fertilizer_info, soil_moisture_thresholds, irrigation_info)
            status     = "✓ Sent" if fb_ok else "FB Error"
            full_recommendation = f"{recommendation}\n\n[Soil Moisture Thresholds Rationale]\n{soil_moisture_rationale}"
            log_gemini(wb, ws_gemini, now, period_str,
                       full_recommendation, json.dumps(actions), status)
            log_schedule(wb, ws_schedule, now, period_str, schedule)

            buffer        = []
            period_start  = now
            next_analysis = now + timedelta(hours=ANALYSIS_HOURS)
            print(f"[Analysis] Done. Next: {next_analysis.strftime('%H:%M:%S')}\n")

        time.sleep(POLL_INTERVAL)

if __name__ == "__main__":
    main()